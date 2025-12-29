"""
🎯 L'ATELIER DE VINCENT - Application de Gestion CA
Application web créée avec Streamlit pour remplacer votre Excel

Auteur : Vincent
Date : Décembre 2024
"""

# ==================== IMPORTS ====================

import streamlit as st
import pandas as pd
import plotly.express as px
from datetime import datetime, timedelta
import os
import calendar
import locale
import time
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image as RLImage, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from io import BytesIO

# Configuration du locale français (avec gestion d'erreur pour Streamlit Cloud)
try:
    locale.setlocale(locale.LC_TIME, "fr_FR.UTF-8")
except locale.Error:
    try:
        locale.setlocale(locale.LC_TIME, "fr_FR")
    except locale.Error:
        try:
            locale.setlocale(locale.LC_TIME, "French_France.1252")
        except locale.Error:
            # Si aucun locale français n'est disponible, on continue sans
            # Les noms de jours/mois sont déjà en français dans le code
            pass

# ==================== CONFIGURATION ====================

st.set_page_config(
    page_title="L'Atelier de Vincent",
    page_icon="assets/logo.png",  # Utilise votre logo comme favicon
    layout="wide",
    initial_sidebar_state="expanded"
)

# Configuration PWA pour utiliser votre logo sur mobile
st.markdown("""
    <head>
        <meta name="application-name" content="L'Atelier de Vincent">
        <meta name="apple-mobile-web-app-title" content="Atelier Vincent">
        <meta name="apple-mobile-web-app-capable" content="yes">
        <meta name="mobile-web-app-capable" content="yes">
        <meta name="theme-color" content="#A89332">
        <link rel="apple-touch-icon" href="assets/logo.png">
        <link rel="icon" type="image/png" sizes="192x192" href="assets/logo.png">
        <link rel="manifest" href="data:application/json;base64,ewogICJuYW1lIjogIkwnQXRlbGllciBkZSBWaW5jZW50IiwKICAic2hvcnRfbmFtZSI6ICJBdGVsaWVyIFZpbmNlbnQiLAogICJkZXNjcmlwdGlvbiI6ICJHZXN0aW9uIENBIHBvdXIgTCdBdGVsaWVyIGRlIFZpbmNlbnQiLAogICJzdGFydF91cmwiOiAiLyIsCiAgImRpc3BsYXkiOiAic3RhbmRhbG9uZSIsCiAgImJhY2tncm91bmRfY29sb3IiOiAiI0Y1RjVGMCIsCiAgInRoZW1lX2NvbG9yIjogIiNBODkzMzIiLAogICJpY29ucyI6IFsKICAgIHsKICAgICAgInNyYyI6ICJhc3NldHMvbG9nby5wbmciLAogICAgICAic2l6ZXMiOiAiNTEyeDUxMiIsCiAgICAgICJ0eXBlIjogImltYWdlL3BuZyIKICAgIH0KICBdCn0=">
    </head>
""", unsafe_allow_html=True)

# ==================== PROTECTION PAR MOT DE PASSE ====================

def verifier_mot_de_passe():
    """Retourne True si le mot de passe est correct."""
    
    if st.session_state.get("password_correct", False):
        return True

    st.title("🔒 L'Atelier de Vincent")
    st.markdown("### Veuillez vous connecter pour accéder à l'application")
    
    password = st.text_input(
        "Mot de passe", 
        type="password",
        placeholder="Entrez le mot de passe"
    )
    
    if st.button("Se connecter", use_container_width=True):
        if password == "3108":
            st.session_state["password_correct"] = True
            st.rerun()
        else:
            st.error("😕 Mot de passe incorrect. Réessayez.")
    
    return False

# ==================== FONCTIONS UTILES ====================

@st.cache_data
def charger_donnees(fichier_excel):
    """Charge les données depuis votre fichier Excel"""
    try:
        df = pd.read_excel(fichier_excel, sheet_name="Données")
        
        df['date'] = pd.to_datetime(df['Date'], errors='coerce')
        df['montant'] = pd.to_numeric(df['Valeur'], errors='coerce')
        df['nb_collaborateurs'] = pd.to_numeric(df['Nb_Collaborateurs'], errors='coerce').fillna(0).astype(int)
        
        df = df.dropna(subset=['date', 'montant'])
        df = df[['date', 'montant', 'nb_collaborateurs']].copy()
        
        return df
    except Exception as e:
        st.error(f"Erreur lors du chargement : {e}")
        return None

def calculer_exercice(date):
    """Calcule l'exercice fiscal (juillet à juin)"""
    if date.month >= 7:
        return f"{date.year}/{date.year + 1}"
    else:
        return f"{date.year - 1}/{date.year}"

def formater_euro(montant):
    """Formate un nombre en euros français"""
    return f"{montant:,.2f} €".replace(",", " ").replace(".", ",")

def generer_pdf_suivi(donnees_tableau, mois_selectionne, annee_mois_n, annee_mois_n_moins_1, total_n, total_n_moins_1, evolution_euro, evolution_pct):
    """Génère un PDF du tableau de suivi mensuel optimisé pour tenir sur une page A4 paysage"""
    buffer = BytesIO()
    
    # Créer le document en mode paysage
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1.5*cm,
        bottomMargin=1*cm
    )
    
    elements = []
    
    # Ajouter le logo en haut
    try:
        logo_path = "assets/logo_noir.png"
        if os.path.exists(logo_path):
            logo = RLImage(logo_path, width=3*cm, height=3*cm)
            elements.append(logo)
            elements.append(Spacer(1, 0.3*cm))
    except:
        pass
    
    # Titre
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.black,
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    title_text = f"Suivi Mensuel - {mois_selectionne} {annee_mois_n} vs {mois_selectionne} {annee_mois_n_moins_1}"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*cm))
    
    # Préparer les données du tableau
    table_data = [['Jour', 'Date N-1', 'Date N', 'Montant N-1', 'Nb C. N-1', 'Montant N', 'Nb C. N']]
    
    for row in donnees_tableau:
        table_data.append([
            row['Jour'][:3],  # Abréger les jours (Lun, Mar, etc.)
            row['Date N-1'],
            row['Date N'],
            row['Montant N-1'],
            row['Nb Collab N-1'],
            row['Montant N'],
            row['Nb Collab N']
        ])
    
    # Créer le tableau avec des largeurs optimisées
    col_widths = [2*cm, 2.5*cm, 2.5*cm, 3*cm, 1.8*cm, 3*cm, 1.8*cm]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Style du tableau
    table.setStyle(TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Corps du tableau
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.4*cm))
    
    # Totaux
    totaux_style = ParagraphStyle(
        'Totaux',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.black,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    totaux_text = f"""
    <b>Total {mois_selectionne} {annee_mois_n_moins_1}:</b> {formater_euro(total_n_moins_1)} | 
    <b>Total {mois_selectionne} {annee_mois_n}:</b> {formater_euro(total_n)} | 
    <b>Évolution:</b> {formater_euro(evolution_euro)} ({evolution_pct:+.1f}%)
    """
    
    totaux = Paragraph(totaux_text, totaux_style)
    elements.append(totaux)
    
    # Pied de page
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    
    date_generation = datetime.now().strftime("%d/%m/%Y à %H:%M")
    footer = Paragraph(f"<i>Document généré le {date_generation} - L'Atelier de Vincent</i>", footer_style)
    elements.append(Spacer(1, 0.3*cm))
    elements.append(footer)
    
    # Construire le PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

def enregistrer_transaction(fichier_excel, date_saisie, montant, nb_collaborateurs):
    """Enregistre une nouvelle transaction dans la feuille Données"""
    try:
        from openpyxl import load_workbook
        from datetime import datetime as dt
        
        # Charger le workbook
        wb = load_workbook(fichier_excel, keep_vba=True)
        ws = wb['Données']
        
        # Préparer les données
        annee = date_saisie.year
        date_str = date_saisie.strftime('%Y-%m-%d')
        cle = f"{annee}|{date_str}"
        
        # Noms des jours et mois en français
        jours_fr = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
        mois_fr = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin', 
                   'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']
        
        jour_semaine = jours_fr[date_saisie.weekday()]
        mois_nom = mois_fr[date_saisie.month - 1]
        
        # Vérifier si la date existe déjà (comparer la DATE, pas la clé)
        ligne_existante = None
        for row in range(2, ws.max_row + 1):
            cell_date = ws.cell(row, 3).value  # Colonne C : Date
            
            # Convertir en datetime si nécessaire
            if cell_date:
                if isinstance(cell_date, str):
                    try:
                        cell_date = dt.strptime(cell_date, '%Y-%m-%d')
                    except:
                        continue
                
                # Comparer uniquement la date (sans l'heure)
                if isinstance(cell_date, dt):
                    if cell_date.date() == date_saisie.date():
                        ligne_existante = row
                        break
        
        if ligne_existante:
            if montant == 0:
                # SUPPRESSION : Montant = 0
                ws.delete_rows(ligne_existante)
                message = f"🗑️ Transaction SUPPRIMÉE pour le {date_saisie.strftime('%d/%m/%Y')}"
            else:
                # MISE À JOUR : La date existe déjà
                ws.cell(row=ligne_existante, column=6, value=montant)
                ws.cell(row=ligne_existante, column=7, value=nb_collaborateurs)
                message = f"✅ Transaction MISE À JOUR : {formater_euro(montant)} le {date_saisie.strftime('%d/%m/%Y')} ({nb_collaborateurs} collaborateur{'s' if nb_collaborateurs > 1 else ''})"
        else:
            if montant == 0:
                # Pas de création si montant = 0 et date inexistante
                message = f"ℹ️ Aucune donnée à supprimer pour le {date_saisie.strftime('%d/%m/%Y')}"
            else:
                # AJOUT : Nouvelle date
                prochaine_ligne = ws.max_row + 1
                
                ws.cell(row=prochaine_ligne, column=1, value=cle)
                ws.cell(row=prochaine_ligne, column=2, value=annee)
                ws.cell(row=prochaine_ligne, column=3, value=date_saisie)
                ws.cell(row=prochaine_ligne, column=4, value=jour_semaine)
                ws.cell(row=prochaine_ligne, column=5, value=mois_nom)
                ws.cell(row=prochaine_ligne, column=6, value=montant)
                ws.cell(row=prochaine_ligne, column=7, value=nb_collaborateurs)
                
                message = f"✅ Transaction AJOUTÉE : {formater_euro(montant)} le {date_saisie.strftime('%d/%m/%Y')} ({nb_collaborateurs} collaborateur{'s' if nb_collaborateurs > 1 else ''})"
        
        # Sauvegarder
        wb.save(fichier_excel)
        wb.close()
        
        return True, message
        
    except Exception as e:
        return False, f"❌ Erreur lors de l'enregistrement : {str(e)}"

# ==================== SIDEBAR ====================

st.sidebar.title("📊 L'Atelier de Vincent")
st.sidebar.markdown("---")

fichier_excel = st.sidebar.text_input(
    "📁 Chemin du fichier Excel",
    value="CA_Atelier_Vincent_B2C2_vers_D4E4.xlsm",
    help="Entrez le chemin complet de votre fichier Excel"
)

page = st.sidebar.radio(
    "Navigation",
    ["🏠 Accueil", "📊 Suivi", "📈 Historique", "➕ Saisie", "⚙️ Données brutes"]
)

st.sidebar.markdown("---")
st.sidebar.info("💡 Application créée pour gérer votre chiffre d'affaires")

# ==================== VÉRIFICATION MOT DE PASSE ====================

if not verifier_mot_de_passe():
    st.stop()

# ==================== CHARGEMENT DES DONNÉES ====================

if os.path.exists(fichier_excel):
    df = charger_donnees(fichier_excel)
    
    if df is not None and not df.empty:
        # Trouver la dernière date avec une valeur > 0
        df_avec_valeur = df[df['montant'] > 0]
        if not df_avec_valeur.empty:
            derniere_date = df_avec_valeur['date'].max()
        else:
            derniere_date = df['date'].max()
        
        # Ajouter colonnes calculées
        df['exercice'] = df['date'].apply(calculer_exercice)
        df['annee'] = df['date'].dt.year
        df['mois'] = df['date'].dt.month
        df['jour_semaine'] = df['date'].dt.day_name()
        
        # ==================== PAGE ACCUEIL ====================

        if page == "🏠 Accueil":
            # En-tête centré
            st.markdown("""
            <h1 style='text-align: center;'>Tableau de Bord<br>L'Atelier de Vincent</h1>
            """, unsafe_allow_html=True)
            
            st.markdown("### 👋 Bonjour Vincent !")
            
            derniere_date_str = derniere_date.strftime("%d/%m/%Y")
            st.markdown(f"### Voici où nous en sommes à la date du : **{derniere_date_str}**")
            st.markdown("---")
            
            # ========== SECTION 1 : JOURNALIER ==========
            st.subheader("📅 Comparaison Journalière")
            
            date_n = derniere_date
            jour_semaine_n = date_n.strftime('%A')
            
            # Trouver le même jour de semaine l'année précédente (avec gestion 29 février)
            try:
                date_n_moins_1_approx = date_n.replace(year=date_n.year - 1)
            except ValueError:
                # Cas du 29 février en année non bissextile → utiliser 28 février
                date_n_moins_1_approx = datetime(date_n.year - 1, 2, 28)
            
            # Chercher le même jour de semaine dans une fenêtre de +/- 3 jours
            for delta in range(-3, 4):
                date_candidate = date_n_moins_1_approx + timedelta(days=delta)
                if date_candidate.strftime('%A') == jour_semaine_n:
                    date_n_moins_1 = date_candidate
                    break
            
            ca_jour_n = df[df['date'] == date_n]['montant'].sum()
            ca_jour_n_moins_1 = df[df['date'] == date_n_moins_1]['montant'].sum()
            
            evolution_jour_euro = ca_jour_n - ca_jour_n_moins_1
            evolution_jour_pct = (evolution_jour_euro / ca_jour_n_moins_1 * 100) if ca_jour_n_moins_1 != 0 else 0
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric(
                   f"CA du {date_n_moins_1.strftime('%d/%m/%Y')}",
   		   formater_euro(ca_jour_n_moins_1),
                   help=f"{jour_semaine_n} {date_n_moins_1.strftime('%d/%m/%Y')}"
               )


            with col2:
                st.metric(
                    f"CA du **{derniere_date_str}**", 
                    formater_euro(ca_jour_n),
                    help=f"{jour_semaine_n} {date_n.strftime('%d/%m/%Y')}"
                )
            with col3:
                st.metric("Évolution €", formater_euro(evolution_jour_euro))
            with col4:
                st.metric("Évolution %", f"{evolution_jour_pct:+.1f}%")
            
            st.markdown("---")
            
            # ========== SECTION 2 : MENSUEL ==========
            st.subheader("📊 Comparaison Mensuelle")
            
            mois_actuel = date_n.month
            annee_actuelle = date_n.year
            jour_actuel = date_n.day
            
            # Cumul mois N
            debut_mois_n = date_n.replace(day=1)
            df_mois_n = df[(df['date'] >= debut_mois_n) & (df['date'] <= date_n)]
            cumul_mois_n = df_mois_n['montant'].sum()
            
            nb_jours_ecoules = jour_actuel
            
            # Cumul mois N-1 : MÊME MOIS, année précédente
            mois_n_moins_1 = mois_actuel
            annee_n_moins_1 = annee_actuelle - 1
            
            debut_mois_n_moins_1 = datetime(annee_n_moins_1, mois_n_moins_1, 1)
            
            dernier_jour_mois_n_moins_1 = calendar.monthrange(annee_n_moins_1, mois_n_moins_1)[1]
            jour_fin_n_moins_1 = min(nb_jours_ecoules, dernier_jour_mois_n_moins_1)
            fin_mois_n_moins_1 = datetime(annee_n_moins_1, mois_n_moins_1, jour_fin_n_moins_1)
            
            df_mois_n_moins_1 = df[(df['date'] >= debut_mois_n_moins_1) & (df['date'] <= fin_mois_n_moins_1)]
            cumul_mois_n_moins_1 = df_mois_n_moins_1['montant'].sum()
            
            evolution_mois_euro = cumul_mois_n - cumul_mois_n_moins_1
            evolution_mois_pct = (evolution_mois_euro / cumul_mois_n_moins_1 * 100) if cumul_mois_n_moins_1 != 0 else 0
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric(
                    "Cumul Mois N-1", 
                    formater_euro(cumul_mois_n_moins_1),
                    help=f"Du 1er au {jour_fin_n_moins_1} {calendar.month_name[mois_n_moins_1]} {annee_n_moins_1} ({jour_fin_n_moins_1} jours)"
                )
            with col2:
                st.metric(
                    "Cumul Mois", 
                    formater_euro(cumul_mois_n),
                    help=f"Du 1er au {jour_actuel} {date_n.strftime('%B %Y')} ({nb_jours_ecoules} jours)"
                )
            with col3:
                st.metric("Évolution €", formater_euro(evolution_mois_euro))
            with col4:
                st.metric("Évolution %", f"{evolution_mois_pct:+.1f}%")
            
            # ========== MESSAGE MOTIVANT ==========
            st.markdown("")
            
            # CA TOTAL du mois de l'année dernière (mois complet)
            debut_mois_complet_n_moins_1 = datetime(annee_n_moins_1, mois_n_moins_1, 1)
            dernier_jour_complet = calendar.monthrange(annee_n_moins_1, mois_n_moins_1)[1]
            fin_mois_complet_n_moins_1 = datetime(annee_n_moins_1, mois_n_moins_1, dernier_jour_complet)
            
            df_mois_complet_n_moins_1 = df[(df['date'] >= debut_mois_complet_n_moins_1) & 
                                            (df['date'] <= fin_mois_complet_n_moins_1)]
            ca_total_mois_n_moins_1 = df_mois_complet_n_moins_1['montant'].sum()
            
            # Reste à faire
            reste_a_faire = ca_total_mois_n_moins_1 - cumul_mois_n
            
            # Nom du mois en français
            mois_fr_noms = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin',
                            'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']
            nom_mois_n_moins_1 = mois_fr_noms[mois_n_moins_1 - 1]
            
            if reste_a_faire > 0:
                st.info(
                    f"🎯 **Objectif :** Pour atteindre le CA de **{nom_mois_n_moins_1} {annee_n_moins_1}** "
                    f"({formater_euro(ca_total_mois_n_moins_1)}), il reste **{formater_euro(reste_a_faire)}** à faire."
                )
            else:
                depassement = abs(reste_a_faire)
                st.success(
                    f"🎉 **Bravo !** Vous avez dépassé le CA de **{nom_mois_n_moins_1} {annee_n_moins_1}** "
                    f"({formater_euro(ca_total_mois_n_moins_1)}) de **{formater_euro(depassement)}** !"
                )
            
            st.markdown("---")

            
            # ========== SECTION 3 : ANNUEL ==========
            st.subheader("📈 Comparaison Annuelle (Exercice)")
            
            exercice_actuel = calculer_exercice(date_n)
            annee_debut_exercice = int(exercice_actuel.split('/')[0])
            
            debut_exercice_n = datetime(annee_debut_exercice, 7, 1)
            df_exercice_n = df[(df['date'] >= debut_exercice_n) & (df['date'] <= date_n)]
            cumul_exercice_n = df_exercice_n['montant'].sum()
            
            # Même période exercice précédent (utilise date_n_moins_1 du calcul journalier)
            debut_exercice_n_moins_1 = datetime(annee_debut_exercice - 1, 7, 1)
            df_exercice_n_moins_1 = df[(df['date'] >= debut_exercice_n_moins_1) & (df['date'] <= date_n_moins_1)]
            cumul_exercice_n_moins_1 = df_exercice_n_moins_1['montant'].sum()
            
            nb_jours_exercice_n = (date_n - debut_exercice_n).days + 1
            nb_jours_exercice_n_moins_1 = (date_n_moins_1 - debut_exercice_n_moins_1).days + 1
            
            evolution_exercice_euro = cumul_exercice_n - cumul_exercice_n_moins_1
            evolution_exercice_pct = (evolution_exercice_euro / cumul_exercice_n_moins_1 * 100) if cumul_exercice_n_moins_1 != 0 else 0
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric(
                    "Cumul Année N-1", 
                    formater_euro(cumul_exercice_n_moins_1),
                    help=f"Du 1er juillet {annee_debut_exercice - 1} au {date_n_moins_1.strftime('%d/%m/%Y')} ({nb_jours_exercice_n_moins_1} jours)"
                )
            with col2:
                st.metric(
                    "Cumul Année N", 
                    formater_euro(cumul_exercice_n),
                    help=f"Du 1er juillet {annee_debut_exercice} au {date_n.strftime('%d/%m/%Y')} ({nb_jours_exercice_n} jours)"
                )
            with col3:
                st.metric("Évolution €", formater_euro(evolution_exercice_euro))
            with col4:
                st.metric("Évolution %", f"{evolution_exercice_pct:+.1f}%")
            
            st.markdown("---")
            
            
           # ========== SECTION 4 : FORMULAIRE DE SAISIE ==========
            st.subheader("➕ Saisir une nouvelle entrée")
            
            with st.form("formulaire_saisie_accueil"):
                st.markdown("**📅 Date**")
                col_jour, col_mois, col_annee = st.columns(3)
                
                # Date du jour par défaut
                aujourd_hui = datetime.now()
                
                with col_jour:
                    jour = st.selectbox(
                        "Jour",
                        options=list(range(1, 32)),
                        index=aujourd_hui.day - 1,
                        label_visibility="collapsed"
                    )
                
                with col_mois:
                    mois_fr = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                               'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
                    mois = st.selectbox(
                        "Mois",
                        options=mois_fr,
                        index=aujourd_hui.month - 1,
                        label_visibility="collapsed"
                    )
                    mois_numero = mois_fr.index(mois) + 1
                
                with col_annee:
                    annee = st.selectbox(
                        "Année",
                        options=list(range(2019, 2031)),
                        index=list(range(2019, 2031)).index(aujourd_hui.year),
                        label_visibility="collapsed"
                    )
                
                # Construire la date
                try:
                    date_saisie = datetime(annee, mois_numero, jour)
                except ValueError:
                    # Si la date est invalide (ex: 31 février)
                    st.error("⚠️ Date invalide")
                    date_saisie = aujourd_hui
                
                st.markdown("**💰 Montant**")
                montant_saisie = st.number_input(
                    "Montant (€)",
                    min_value=0.0,
                    value=0.0,
                    step=0.01,
                    format="%.2f",
                    label_visibility="collapsed"
                )
                
                st.markdown("**👥 Nombre de collaborateurs**")
                nb_collaborateurs = st.selectbox(
                    "Nombre de collaborateurs",
                    options=[1, 2, 3, 4],
                    index=1,  # Par défaut : 2 personnes (Patron + CDI)
                    label_visibility="collapsed",
                    help="1 = Patron seul | 2 = Patron + CDI | 3 = Patron + CDI + Stagiaire | 4 = Patron + CDI + 2 Stagiaires"
                )
                
                submit = st.form_submit_button("✅ Enregistrer", use_container_width=True)
                
                if submit:
                    if montant_saisie >= 0:
                        succes, message = enregistrer_transaction(fichier_excel, date_saisie, montant_saisie, nb_collaborateurs)
                        
                        if succes:
                            st.success(message)
                            st.balloons()
                            st.cache_data.clear()
                            st.rerun()
                        else:
                            st.error(message)
                            
            st.markdown("---")



        # ==================== AUTRES PAGES ====================
        
        elif page == "📊 Suivi":
            st.title("📊 Suivi Mensuel par Exercice")
        
            # ========== SÉLECTION DE L'EXERCICE ==========
            exercices_disponibles = []
            annees = sorted(df['date'].dt.year.unique())
        
            for annee in annees:
                exercices_disponibles.append(f"{annee}/{annee + 1}")
        
            # Retirer les doublons et trier
            exercices_disponibles = sorted(list(set(exercices_disponibles)))
        
            # Exercice actuel par défaut
            date_actuelle = datetime.now()
            exercice_actuel = calculer_exercice(date_actuelle)
        
            if exercice_actuel in exercices_disponibles:
                index_defaut = exercices_disponibles.index(exercice_actuel)
            else:
                index_defaut = len(exercices_disponibles) - 1
        
            col1, col2 = st.columns([2, 3])
        
            with col1:
                exercice_selectionne = st.selectbox(
                    "📅 Choisir l'exercice",
                    options=exercices_disponibles,
                    index=index_defaut
                )
        
            with col2:
                mois_liste = ['Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre',
                              'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin']
            
                # Mois actuel par défaut
                mois_actuel_index = (date_actuelle.month - 7) % 12
            
                mois_selectionne = st.selectbox(
                    "📆 Choisir le mois",
                    options=mois_liste,
                    index=mois_actuel_index
                )
        
            st.markdown("---")
        
            # ========== CALCUL DES DATES ==========
            annee_debut_exercice = int(exercice_selectionne.split('/')[0])
            
            # Mapper le nom du mois à son vrai numéro (1-12)
            mois_mapping = {
                'Juillet': 7, 'Août': 8, 'Septembre': 9, 'Octobre': 10, 'Novembre': 11, 'Décembre': 12,
                'Janvier': 1, 'Février': 2, 'Mars': 3, 'Avril': 4, 'Mai': 5, 'Juin': 6
            }
            mois_numero = mois_mapping[mois_selectionne]
        
            # Ajuster l'année du mois selon l'exercice
            if mois_numero >= 7:  # Juillet à Décembre
                annee_mois_n = annee_debut_exercice
            else:  # Janvier à Juin
                annee_mois_n = annee_debut_exercice + 1
        
            # Calculer l'année N-1
            annee_mois_n_moins_1 = annee_mois_n - 1
        
            # Nombre de jours dans le mois
            nb_jours_mois = calendar.monthrange(annee_mois_n, mois_numero)[1]
        
            # ========== CRÉATION DU TABLEAU ==========
            st.subheader(f"📋 {mois_selectionne} {annee_mois_n} vs {mois_selectionne} {annee_mois_n_moins_1}")
            
            # Bouton Export PDF (sera activé après calcul des données)
            placeholder_pdf_button = st.empty()
        
            # Créer les données du tableau
            donnees_tableau = []
        
            for jour in range(1, nb_jours_mois + 1):
                date_n = datetime(annee_mois_n, mois_numero, jour)
                jour_semaine = date_n.weekday()  # 0 = Lundi, 6 = Dimanche
            
                # Trouver la date N-1 correspondante (même jour de la semaine)
                # Chercher le même jour de la semaine dans l'année N-1
                date_reference_n_moins_1 = datetime(annee_mois_n_moins_1, mois_numero, jour)
                jours_diff = (jour_semaine - date_reference_n_moins_1.weekday()) % 7
            
                if jours_diff <= 3:
                    date_n_moins_1 = date_reference_n_moins_1 + timedelta(days=jours_diff)
                else:
                    date_n_moins_1 = date_reference_n_moins_1 - timedelta(days=7 - jours_diff)
            
                # Récupérer les montants et le nombre de collaborateurs
                data_n = df[df['date'] == date_n]
                montant_n = data_n['montant'].sum()
                nb_collab_n = data_n['nb_collaborateurs'].max() if not data_n.empty else 0
                
                data_n_moins_1 = df[df['date'] == date_n_moins_1]
                montant_n_moins_1 = data_n_moins_1['montant'].sum()
                nb_collab_n_moins_1 = data_n_moins_1['nb_collaborateurs'].max() if not data_n_moins_1.empty else 0
            
                # Noms des jours en français
                jours_fr = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
                nom_jour = jours_fr[jour_semaine]
            
                donnees_tableau.append({
                    'Jour': nom_jour,
                    'Date N-1': date_n_moins_1.strftime('%d/%m/%Y'),
                    'Date N': date_n.strftime('%d/%m/%Y'),
                    'Montant N-1': formater_euro(montant_n_moins_1) if montant_n_moins_1 > 0 else '-',
                    'Nb Collab N-1': str(nb_collab_n_moins_1) if montant_n_moins_1 > 0 else '-',
                    'Montant N': formater_euro(montant_n) if montant_n > 0 else '-',
                    'Nb Collab N': str(nb_collab_n) if montant_n > 0 else '-'
                })
        
            # Créer le DataFrame
            df_tableau = pd.DataFrame(donnees_tableau)
            
            # Calculer les totaux pour le PDF (avant l'affichage)
            debut_mois_n = datetime(annee_mois_n, mois_numero, 1)
            fin_mois_n = datetime(annee_mois_n, mois_numero, nb_jours_mois)
            df_mois_n = df[(df['date'] >= debut_mois_n) & (df['date'] <= fin_mois_n)]
            total_n = df_mois_n['montant'].sum()
            
            debut_mois_n_moins_1 = datetime(annee_mois_n_moins_1, mois_numero, 1)
            fin_mois_n_moins_1 = datetime(annee_mois_n_moins_1, mois_numero, nb_jours_mois)
            df_mois_n_moins_1 = df[(df['date'] >= debut_mois_n_moins_1) & (df['date'] <= fin_mois_n_moins_1)]
            total_n_moins_1 = df_mois_n_moins_1['montant'].sum()
            
            evolution_euro = total_n - total_n_moins_1
            evolution_pct = (evolution_euro / total_n_moins_1 * 100) if total_n_moins_1 != 0 else 0
            
            # Bouton Export PDF avec le placeholder
            with placeholder_pdf_button:
                pdf_buffer = generer_pdf_suivi(
                    donnees_tableau, 
                    mois_selectionne, 
                    annee_mois_n, 
                    annee_mois_n_moins_1,
                    total_n,
                    total_n_moins_1,
                    evolution_euro,
                    evolution_pct
                )
                
                st.download_button(
                    label="📄 Exporter en PDF",
                    data=pdf_buffer,
                    file_name=f"Suivi_{mois_selectionne}_{annee_mois_n}.pdf",
                    mime="application/pdf",
                    use_container_width=False
                )
            
            st.markdown("---")
        
            # Afficher le tableau
            st.dataframe(
                df_tableau,
                hide_index=True,
                use_container_width=True,
                height=600,
                column_config={
                    "Jour": st.column_config.TextColumn("Jour", width="small"),
                    "Date N-1": st.column_config.TextColumn("Date N-1", width="medium"),
                    "Date N": st.column_config.TextColumn("Date N", width="medium"),
                    "Montant N-1": st.column_config.TextColumn("Montant N-1", width="medium"),
                    "Nb Collab N-1": st.column_config.TextColumn("Nb Collab N-1", width="small"),
                    "Montant N": st.column_config.TextColumn("Montant N", width="medium"),
                    "Nb Collab N": st.column_config.TextColumn("Nb Collab N", width="small")
                }
            )
        
            # ========== TOTAUX ==========
            st.markdown("---")
        
            col1, col2, col3, col4 = st.columns(4)
        
            with col1:
                st.metric(
                    f"Total {mois_selectionne} {annee_mois_n_moins_1}",
                    formater_euro(total_n_moins_1)
                )
        
            with col2:
                st.metric(
                    f"Total {mois_selectionne} {annee_mois_n}",
                    formater_euro(total_n)
                )
        
            with col3:
                st.metric("Évolution €", formater_euro(evolution_euro))
        
            with col4:
                st.metric("Évolution %", f"{evolution_pct:+.1f}%")
        
        elif page == "📈 Historique":
            st.title("📈 Historique")
            st.info("Page Historique en construction")
        
        elif page == "➕ Saisie":
            st.title("➕ Saisie de données")
            st.info("Utilisez le formulaire sur la page d'accueil")
        
        elif page == "⚙️ Données brutes":
            st.title("⚙️ Données brutes")
            st.dataframe(df, use_container_width=True)
    
    else:
        st.error("❌ Impossible de charger les données du fichier Excel")
else:
    st.error(f"❌ Le fichier '{fichier_excel}' n'existe pas. Vérifiez le chemin dans la sidebar.")
